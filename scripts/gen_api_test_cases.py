# -*- coding: utf-8 -*-
"""
生成《接口扁平化测试用例》Excel（.xlsx）。

扁平化测试目标：对系统全部 HTTP 接口逐个直接测试，覆盖每个接口的
正常、边界、异常、业务规则与归属/权限分支，形成全量用例矩阵。

运行：
    python3 scripts/gen_api_test_cases.py
产物：
    docs/test/接口扁平化测试用例.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------- 列定义（13 列） -----------------------------
# 用例数据每个 tuple 为 11 字段：(id, prio, type, pre, method, path, headers, body, status, expect, note)
# 输出时在 id 后插入“模块”“接口”两列，共 13 列。
COLS = [
    ("用例编号", 15), ("模块", 15), ("接口", 30), ("优先级", 8), ("测试类型", 11),
    ("前置条件", 24), ("请求方法", 9), ("请求路径", 34), ("请求头", 24),
    ("请求参数 / 请求体", 48), ("预期状态码", 11), ("预期结果", 54), ("备注", 26),
]
# 左对齐（长文本、自动换行）的列号
LEFT_COLS = {6, 9, 10, 12, 13}  # 前置条件 / 请求头 / 请求参数 / 预期结果 / 备注

MODULE_FILL = {
    "Excel 导入导出": "1F4E79",
    "异步任务中心": "375623",
    "报表运行控制": "7F6000",
    "文件上传中心": "833C00",
}
PRIO_FILL = {"P0": "C00000", "P1": "BF8F00", "P2": "548235"}

# ----------------------------- 用例数据 -----------------------------
CASES = []

def add(module, endpoint, cases):
    for c in cases:
        # c = (id, prio, type, pre, method, path, headers, body, status, expect, note)
        CASES.append((c[0], module, endpoint, c[1], c[2], c[3], c[4],
                      c[5], c[6], c[7], c[8], c[9], c[10]))

# ============================ Excel 导入导出 /api/excel ============================
EXC = "Excel 导入导出"
add(EXC, "GET /api/excel/count", [
    ("EXC-COUNT-01", "P0", "正常", "系统已启动", "GET", "/api/excel/count", "", "",
     "200", "返回 JSON，含 count 字段（当前学生总数，可为 0）", "无鉴权要求"),
    ("EXC-COUNT-02", "P2", "边界", "学生表为空", "GET", "/api/excel/count", "", "",
     "200", "count=0", "首次启动空库场景"),
])
add(EXC, "POST /api/excel/seed/{count}", [
    ("EXC-SEED-01", "P0", "正常", "无", "POST", "/api/excel/seed/5", "", "",
     "200", "返回 inserted、count、elapsedMs；count 较前增加 inserted", ""),
    ("EXC-SEED-02", "P2", "边界", "无", "POST", "/api/excel/seed/0", "", "",
     "200", "inserted=0，count 不变", "count=0 不报错"),
    ("EXC-SEED-03", "P2", "异常", "无", "POST", "/api/excel/seed/abc", "", "",
     "400", "路径变量非整数，Spring 类型转换失败 → 400", ""),
])
add(EXC, "POST /api/excel/export", [
    ("EXC-EXP-01", "P0", "正常", "已存在学生数据", "POST", "/api/excel/export", "X-User-Id: user-a", "",
     "200", "返回 ExportTaskResponse，含 taskId，status 为 QUEUED/RUNNING", "异步任务"),
    ("EXC-EXP-02", "P1", "正常", "不传归属头", "POST", "/api/excel/export", "", "",
     "200", "归属为默认 anonymous，仍可提交，返回 taskId", "默认归属"),
    ("EXC-EXP-03", "P2", "业务规则", "导出队列已满且拒绝策略=abort", "POST", "/api/excel/export", "", "",
     "200", "新任务 status=FAILED，errorMessage 含“导出任务提交失败”", "需先填满队列"),
])
add(EXC, "GET /api/excel/export/{taskId}", [
    ("EXC-EXPS-01", "P0", "正常", "已提交导出任务", "GET", "/api/excel/export/{taskId}", "X-User-Id: user-a", "",
     "200", "返回任务详情；轮询可见 QUEUED→RUNNING→SUCCESS；SUCCESS 时 exported=total、sheetCount=1", ""),
    ("EXC-EXPS-02", "P1", "异常", "taskId 不存在", "GET", "/api/excel/export/not-exist", "X-User-Id: user-a", "",
     "404", "“导出任务不存在”", ""),
    ("EXC-EXPS-03", "P1", "权限", "用其他用户查询他人任务", "GET", "/api/excel/export/{taskId}", "X-User-Id: user-b", "",
     "404", "归属不匹配返回 404，不泄露任务存在性", "越权防探测"),
])
add(EXC, "GET /api/excel/export/{taskId}/download", [
    ("EXC-EXPD-01", "P0", "正常", "任务 status=SUCCESS", "GET", "/api/excel/export/{taskId}/download", "X-User-Id: user-a", "",
     "302", "返回 302 + Location（MinIO 签名下载地址）", ""),
    ("EXC-EXPD-02", "P1", "业务规则", "任务尚未 SUCCESS（QUEUED/RUNNING）", "GET", "/api/excel/export/{taskId}/download", "X-User-Id: user-a", "",
     "409", "“导出任务尚未完成”", ""),
    ("EXC-EXPD-03", "P1", "异常", "任务成功但导出对象缺失", "GET", "/api/excel/export/{taskId}/download", "X-User-Id: user-a", "",
     "404", "“导出文件不存在”", "对象被生命周期清理"),
    ("EXC-EXPD-04", "P1", "权限", "非归属用户下载", "GET", "/api/excel/export/{taskId}/download", "X-User-Id: user-b", "",
     "404", "归属校验先于状态校验，返回 404", ""),
])
add(EXC, "GET /api/excel/template", [
    ("EXC-TPL-01", "P0", "正常", "无", "GET", "/api/excel/template", "", "",
     "200", "Content-Type 为 xlsx；Content-Disposition=attachment；可直接作为导入模板", ""),
    ("EXC-TPL-02", "P2", "边界", "无", "GET", "/api/excel/template", "", "",
     "200", "文件名编码为 UTF-8 百分号编码，中文不乱码", "文件名编码"),
])
add(EXC, "POST /api/excel/import", [
    ("EXC-IMP-01", "P0", "正常", "准备好合法 Excel（含新学号）", "POST", "/api/excel/import", "X-User-Id: user-a",
     "multipart/form-data; file=@students.xlsx", "200",
     "返回 ImportTaskResponse（含 taskId）；完成后 imported=数据行数、batchCount>0", "异步导入"),
    ("EXC-IMP-02", "P0", "正常", "下载模板并填入数据", "POST", "/api/excel/import", "", "file=@student-import-template.xlsx",
     "200", "imported 为数据行数；DB 中存在对应学号", ""),
    ("EXC-IMP-03", "P1", "异常", "未携带 file 参数", "POST", "/api/excel/import", "", "（无 file）",
     "400", "required request part 'file' is not present → 400", ""),
    ("EXC-IMP-04", "P1", "异常", "上传空文件", "POST", "/api/excel/import", "", "file=@empty.xlsx",
     "400", "“上传文件不能为空”", "Controller/Service 校验"),
    ("EXC-IMP-05", "P1", "异常", "损坏/非 Excel 文件", "POST", "/api/excel/import", "", "file=@not-excel.txt",
     "200→任务 FAILED", "提交成功返回 taskId；异步解析失败，任务 status=FAILED", "解析在异步阶段"),
    ("EXC-IMP-06", "P2", "边界", "空模板（仅表头）", "POST", "/api/excel/import", "", "file=@empty-template.xlsx",
     "200", "imported=0、batchCount=0，任务成功", ""),
    ("EXC-IMP-07", "P1", "业务规则", "同一文件出现重复 student_no", "POST", "/api/excel/import", "", "file=@dup-stdno.xlsx",
     "200→任务 FAILED", "导入任务 FAILED；正式表不发生本次导入变更", "原子性"),
    ("EXC-IMP-08", "P2", "业务规则", "回导：导出后改字段保持学号不变", "POST", "/api/excel/import", "", "file=@reimport.xlsx",
     "200", "正式表对应学号记录被更新，总数不增加", "upsert by student_no"),
    ("EXC-IMP-09", "P2", "边界", "文件超过 multipart 限制 200MB", "POST", "/api/excel/import", "", "file=@huge.xlsx (>200MB)",
     "413", "MaxUploadSizeExceededException → 413", ""),
])
add(EXC, "GET /api/excel/import/{taskId}", [
    ("EXC-IMPS-01", "P0", "正常", "已提交导入任务", "GET", "/api/excel/import/{taskId}", "X-User-Id: user-a", "",
     "200", "返回 ImportTaskResponse；进度 completed/total；失败时含 errorMessage 与错误文件信息", ""),
    ("EXC-IMPS-02", "P1", "异常", "taskId 不存在", "GET", "/api/excel/import/not-exist", "X-User-Id: user-a", "",
     "404", "“导入任务不存在”", ""),
    ("EXC-IMPS-03", "P1", "权限", "非归属用户查询", "GET", "/api/excel/import/{taskId}", "X-User-Id: user-b", "",
     "404", "归属不匹配返回 404", ""),
])
add(EXC, "GET /api/excel/import/{taskId}/error-file", [
    ("EXC-ERRF-01", "P1", "正常", "导入存在校验失败行", "GET", "/api/excel/import/{taskId}/error-file", "X-User-Id: user-a", "",
     "302", "返回 302 + Location（错误明细 Excel 签名地址）", ""),
    ("EXC-ERRF-02", "P1", "异常", "导入无错误（无错误文件）", "GET", "/api/excel/import/{taskId}/error-file", "X-User-Id: user-a", "",
     "404", "“导入错误明细文件不存在”", ""),
    ("EXC-ERRF-03", "P1", "权限", "非归属用户", "GET", "/api/excel/import/{taskId}/error-file", "X-User-Id: user-b", "",
     "404", "归属校验返回 404", ""),
])

# ============================ 异步任务中心 /api/tasks ============================
TSK = "异步任务中心"
add(TSK, "GET /api/tasks/{taskId}", [
    ("TSK-GET-01", "P0", "正常", "存在归属任务", "GET", "/api/tasks/{taskId}", "X-User-Id: user-a", "",
     "200", "返回 AsyncTaskResponse（taskId/type/status/progress/retryCount 等）", ""),
    ("TSK-GET-02", "P1", "异常", "taskId 不存在", "GET", "/api/tasks/not-exist", "X-User-Id: user-a", "",
     "404", "“任务不存在”", ""),
    ("TSK-GET-03", "P1", "权限", "查询他人任务", "GET", "/api/tasks/{taskId}", "X-User-Id: user-b", "",
     "404", "归属不匹配返回 404", ""),
])
add(TSK, "POST /api/tasks/page", [
    ("TSK-PAGE-01", "P0", "正常", "存在任务", "POST", "/api/tasks/page", "X-User-Id: user-a",
     '{"pageNo":1,"pageSize":10}', "200", "返回 total/pageNo/pageSize/records；仅返回当前归属任务", ""),
    ("TSK-PAGE-02", "P2", "边界", "无", "POST", "/api/tasks/page", "", "{}",
     "200", "pageNo 缺省=1，pageSize 缺省=20", "缺省分页"),
    ("TSK-PAGE-03", "P2", "边界", "无", "POST", "/api/tasks/page", "X-User-Id: user-a",
     '{"pageNo":0,"pageSize":999}', "200", "pageNo<1 → 1；pageSize 受 max-page-size(100) 限制", ""),
    ("TSK-PAGE-04", "P1", "正常", "存在不同类型/状态任务", "POST", "/api/tasks/page", "X-User-Id: user-a",
     '{"taskType":"IMPORT","status":"SUCCESS"}', "200", "按 taskType 与 status 过滤", ""),
    ("TSK-PAGE-05", "P2", "异常", "无", "POST", "/api/tasks/page", "X-User-Id: user-a",
     '{"status":"GARBAGE"}', "400", "统一异常处理将非法枚举转为 400 COMMON_PARAM_ERROR（2026-08-14 已验证）", "已由 TODO#8 修复"),
    ("TSK-PAGE-06", "P2", "异常", "无", "POST", "/api/tasks/page", "X-User-Id: user-a",
     '{"taskType":"GARBAGE"}', "400", "非法任务类型枚举→400 COMMON_PARAM_ERROR", "已由 TODO#8 修复"),
])
add(TSK, "POST /api/tasks/{taskId}/cancel", [
    ("TSK-CANCEL-01", "P1", "正常", "任务处于 CREATED/RUNNING", "POST", "/api/tasks/{taskId}/cancel", "X-User-Id: user-a", "",
     "200", "canceled=true，任务变为 CANCELED", ""),
    ("TSK-CANCEL-02", "P2", "业务规则", "任务已是 CANCELED", "POST", "/api/tasks/{taskId}/cancel", "X-User-Id: user-a", "",
     "200", "canceled=true（终态 CANCELED 返回 true）", ""),
    ("TSK-CANCEL-03", "P2", "业务规则", "任务已 SUCCESS/FAILED/EXPIRED", "POST", "/api/tasks/{taskId}/cancel", "X-User-Id: user-a", "",
     "200", "canceled=false（终态非 CANCELED 不再变更）", ""),
    ("TSK-CANCEL-04", "P1", "异常", "taskId 不存在", "POST", "/api/tasks/not-exist/cancel", "X-User-Id: user-a", "",
     "404", "“任务不存在”", ""),
    ("TSK-CANCEL-05", "P1", "权限", "取消他人任务", "POST", "/api/tasks/{taskId}/cancel", "X-User-Id: user-b", "",
     "404", "归属校验返回 404", ""),
])
add(TSK, "POST /api/tasks/{taskId}/retry", [
    ("TSK-RETRY-01", "P1", "正常", "任务 FAILED 且 retryCount<max", "POST", "/api/tasks/{taskId}/retry", "X-User-Id: user-a", "",
     "200", "任务重置为 CREATED，retryCount+1", ""),
    ("TSK-RETRY-02", "P2", "业务规则", "任务 CANCELED/EXPIRED", "POST", "/api/tasks/{taskId}/retry", "X-User-Id: user-a", "",
     "200", "允许重试，重置为 CREATED", ""),
    ("TSK-RETRY-03", "P1", "业务规则", "任务 SUCCESS/RUNNING/CREATED", "POST", "/api/tasks/{taskId}/retry", "X-User-Id: user-a", "",
     "409", "“当前任务状态不允许重试”", ""),
    ("TSK-RETRY-04", "P1", "业务规则", "retryCount>=maxRetryCount", "POST", "/api/tasks/{taskId}/retry", "X-User-Id: user-a", "",
     "409", "“任务重试次数已达上限”", ""),
    ("TSK-RETRY-05", "P2", "业务规则", "任务类型无重试处理器", "POST", "/api/tasks/{taskId}/retry", "X-User-Id: user-a", "",
     "409", "“当前任务类型暂不支持重试”", ""),
    ("TSK-RETRY-06", "P1", "异常", "taskId 不存在", "POST", "/api/tasks/not-exist/retry", "X-User-Id: user-a", "",
     "404", "“任务不存在”", ""),
    ("TSK-RETRY-07", "P1", "权限", "重试他人任务", "POST", "/api/tasks/{taskId}/retry", "X-User-Id: user-b", "",
     "404", "归属校验返回 404", ""),
])

# ============================ 报表运行控制 /api/report/student-runs ============================
RPT = "报表运行控制"
add(RPT, "POST /api/report/student-runs/page", [
    ("RPT-PAGE-01", "P0", "正常", "存在运行控制", "POST", "/api/report/student-runs/page", "X-User-Id: user-a",
     '{"pageNo":1,"pageSize":10}', "200", "返回当前归属运行控制列表", ""),
    ("RPT-PAGE-02", "P2", "正常", "按名称过滤", "POST", "/api/report/student-runs/page", "X-User-Id: user-a",
     '{"runName":"学期"}', "200", "按 runName 模糊过滤", ""),
    ("RPT-PAGE-03", "P2", "边界", "无", "POST", "/api/report/student-runs/page", "X-User-Id: user-a",
     '{"pageNo":0,"pageSize":999}', "200", "pageNo=1；pageSize 受 100 限制", ""),
    ("RPT-PAGE-04", "P2", "异常", "无", "POST", "/api/report/student-runs/page", "X-User-Id: user-a",
     '{"status":"GARBAGE"}', "400", "非法状态枚举→400 COMMON_PARAM_ERROR", "已由 TODO#8 修复"),
])
add(RPT, "POST /api/report/student-runs/create", [
    ("RPT-CRT-01", "P0", "正常", "无", "POST", "/api/report/student-runs/create", "X-User-Id: user-a",
     '{"runControlCode":"R001","runName":"2026春季","minAge":18,"maxAge":25}', "200",
     "返回运行控制详情，status=NORMAL，生成 runId", ""),
    ("RPT-CRT-02", "P1", "异常", "无", "POST", "/api/report/student-runs/create", "X-User-Id: user-a",
     '{}', "400", "“运行控制编码不能为空/运行控制名称不能为空”", "缺必填"),
    ("RPT-CRT-03", "P1", "异常", "已存在编码 R001", "POST", "/api/report/student-runs/create", "X-User-Id: user-a",
     '{"runControlCode":"R001","runName":"dup"}', "400", "“运行控制编码已存在”", "唯一约束"),
    ("RPT-CRT-04", "P2", "边界", "无", "POST", "/api/report/student-runs/create", "X-User-Id: user-a",
     '{"runControlCode":"R002","runName":"x","minAge":-1}', "400", "“最小年龄必须在0到150之间”", ""),
    ("RPT-CRT-05", "P2", "边界", "无", "POST", "/api/report/student-runs/create", "X-User-Id: user-a",
     '{"runControlCode":"R003","runName":"x","maxAge":200}', "400", "“最大年龄必须在0到150之间”", ""),
    ("RPT-CRT-06", "P2", "边界", "无", "POST", "/api/report/student-runs/create", "X-User-Id: user-a",
     '{"runControlCode":"R004","runName":"x","minAge":30,"maxAge":20}', "400", "“最小年龄不能大于最大年龄”", ""),
    ("RPT-CRT-07", "P2", "边界", "无", "POST", "/api/report/student-runs/create", "X-User-Id: user-a",
     '{"runControlCode":"R005","runName":"x"}', "200", "可选条件全空也可创建", "仅必填项"),
])
add(RPT, "GET /api/report/student-runs/{runId}", [
    ("RPT-GET-01", "P0", "正常", "存在归属运行控制", "GET", "/api/report/student-runs/{runId}", "X-User-Id: user-a", "",
     "200", "返回运行控制详情", ""),
    ("RPT-GET-02", "P1", "异常", "runId 不存在", "GET", "/api/report/student-runs/not-exist", "X-User-Id: user-a", "",
     "404", "“运行控制不存在”", ""),
    ("RPT-GET-03", "P1", "权限", "查询他人运行控制", "GET", "/api/report/student-runs/{runId}", "X-User-Id: user-b", "",
     "404", "归属不匹配返回 404", ""),
    ("RPT-GET-04", "P2", "业务规则", "运行控制已逻辑删除", "GET", "/api/report/student-runs/{runId}", "X-User-Id: user-a", "",
     "404", "已删除对当前用户不可见", ""),
])
add(RPT, "POST /api/report/student-runs/{runId}/update", [
    ("RPT-UPD-01", "P0", "正常", "存在归属运行控制", "POST", "/api/report/student-runs/{runId}/update", "X-User-Id: user-a",
     '{"runControlCode":"R001","runName":"2026春季改","className":"一班"}', "200", "返回更新后详情", ""),
    ("RPT-UPD-02", "P2", "正常", "改为自身原编码", "POST", "/api/report/student-runs/{runId}/update", "X-User-Id: user-a",
     '{"runControlCode":"R001","runName":"x"}', "200", "编码与自身相同不判重，更新成功", ""),
    ("RPT-UPD-03", "P1", "业务规则", "编码改为他人已用编码", "POST", "/api/report/student-runs/{runId}/update", "X-User-Id: user-a",
     '{"runControlCode":"R002","runName":"x"}', "400", "“运行控制编码已存在”", ""),
    ("RPT-UPD-04", "P1", "异常", "runId 不存在", "POST", "/api/report/student-runs/not-exist/update", "X-User-Id: user-a",
     '{"runControlCode":"R","runName":"x"}', "404", "“运行控制不存在”", ""),
    ("RPT-UPD-05", "P2", "边界", "存在归属运行控制", "POST", "/api/report/student-runs/{runId}/update", "X-User-Id: user-a",
     '{"runControlCode":"R001","runName":"x","minAge":40,"maxAge":30}', "400", "“最小年龄不能大于最大年龄”", ""),
])
add(RPT, "POST /api/report/student-runs/{runId}/delete", [
    ("RPT-DEL-01", "P0", "正常", "存在归属运行控制", "POST", "/api/report/student-runs/{runId}/delete", "X-User-Id: user-a", "",
     "200", "deleted=true；逻辑删除，deleted 写入本行 id", ""),
    ("RPT-DEL-02", "P1", "异常", "runId 不存在", "POST", "/api/report/student-runs/not-exist/delete", "X-User-Id: user-a", "",
     "404", "“运行控制不存在”", ""),
    ("RPT-DEL-03", "P1", "权限", "删除他人运行控制", "POST", "/api/report/student-runs/{runId}/delete", "X-User-Id: user-b", "",
     "404", "归属校验返回 404", ""),
    ("RPT-DEL-04", "P2", "业务规则", "删除后重建同编码", "POST", "/api/report/student-runs/create", "X-User-Id: user-a",
     '{"runControlCode":"R001","runName":"重建"}', "200", "删除后同用户可重建相同编码", ""),
])
add(RPT, "POST /api/report/student-runs/{runId}/run", [
    ("RPT-RUN-01", "P0", "正常", "存在归属运行控制", "POST", "/api/report/student-runs/{runId}/run", "X-User-Id: user-a", "",
     "200", "创建导出任务，返回 ExportTaskResponse；任务 businessKey=runId", ""),
    ("RPT-RUN-02", "P1", "异常", "runId 不存在", "POST", "/api/report/student-runs/not-exist/run", "X-User-Id: user-a", "",
     "404", "“运行控制不存在”", ""),
    ("RPT-RUN-03", "P1", "权限", "运行他人运行控制", "POST", "/api/report/student-runs/{runId}/run", "X-User-Id: user-b", "",
     "404", "归属校验返回 404", ""),
])
add(RPT, "POST /api/report/student-runs/{runId}/tasks", [
    ("RPT-TASKS-01", "P1", "正常", "该运行控制已 run 过", "POST", "/api/report/student-runs/{runId}/tasks", "X-User-Id: user-a",
     '{"pageNo":1,"pageSize":10}', "200", "按 businessKey=runId 分页返回历史导出任务", ""),
    ("RPT-TASKS-02", "P1", "异常", "runId 不存在", "POST", "/api/report/student-runs/not-exist/tasks", "X-User-Id: user-a",
     '{}', "404", "“运行控制不存在”", ""),
    ("RPT-TASKS-03", "P2", "权限", "查询他人运行控制任务", "POST", "/api/report/student-runs/{runId}/tasks", "X-User-Id: user-b",
     '{}', "404", "归属校验返回 404", ""),
])

# ============================ 文件上传中心 /api/files ============================
FIL = "文件上传中心"
add(FIL, "POST /api/files/upload", [
    ("FIL-UP-01", "P0", "正常", "无", "POST", "/api/files/upload", "",
     "multipart/form-data; file=@demo.txt", "200", "返回 FileUploadResponse（fileId/originalName/fileSize/fileMd5/elapsedMs）", ""),
    ("FIL-UP-02", "P1", "异常", "未携带 file", "POST", "/api/files/upload", "", "（无 file）",
     "400", "required request part 'file' is not present → 400", ""),
    ("FIL-UP-03", "P1", "异常", "上传空文件", "POST", "/api/files/upload", "", "file=@empty.txt",
     "400", "“上传文件不能为空”", ""),
])
add(FIL, "POST /api/files/instant-check", [
    ("FIL-IC-01", "P0", "正常", "已上传某文件得到 md5/size", "POST", "/api/files/instant-check", "",
     '{"fileMd5":"<32位md5>","fileSize":123}', "200", "命中：exists=true 并返回 file 信息", "秒传命中"),
    ("FIL-IC-02", "P1", "正常", "md5/size 未命中", "POST", "/api/files/instant-check", "",
     '{"fileMd5":"0000...0001","fileSize":123}', "200", "exists=false，file=null", "秒传未命中"),
    ("FIL-IC-03", "P1", "异常", "无", "POST", "/api/files/instant-check", "",
     '{"fileMd5":"","fileSize":123}', "400", "“文件 MD5 不能为空”", ""),
    ("FIL-IC-04", "P2", "边界", "无", "POST", "/api/files/instant-check", "",
     '{"fileMd5":"not-hex","fileSize":123}', "400", "“文件 MD5 格式不正确”（需 32 位小写十六进制）", ""),
    ("FIL-IC-05", "P2", "边界", "无", "POST", "/api/files/instant-check", "",
     '{"fileMd5":"<32位md5>","fileSize":0}', "400", "“文件大小必须大于 0”", ""),
])
add(FIL, "POST /api/files/direct/init", [
    ("FIL-DI-01", "P0", "正常", "md5 未命中", "POST", "/api/files/direct/init", "",
     '{"originalName":"demo.zip","contentType":"application/zip","fileSize":1024,"fileMd5":"<32位>"}',
     "200", "instant=false；返回 uploadId/fileId/uploadUrl/objectKey/expireMinutes", ""),
    ("FIL-DI-02", "P1", "正常", "md5 命中", "POST", "/api/files/direct/init", "",
     '{"originalName":"demo.zip","fileSize":<size>,"fileMd5":"<已存在md5>"}',
     "200", "instant=true；返回 fileId/file，无需上传", ""),
    ("FIL-DI-03", "P2", "边界", "无", "POST", "/api/files/direct/init", "",
     '{"originalName":"","fileSize":1024,"fileMd5":"<32位>"}',
     "200", "originalName 缺省归一化为 unknown", ""),
    ("FIL-DI-04", "P1", "异常", "无", "POST", "/api/files/direct/init", "",
     '{"originalName":"x","fileSize":1024,"fileMd5":"bad"}',
     "400", "MD5 格式校验失败", ""),
    ("FIL-DI-05", "P2", "边界", "无", "POST", "/api/files/direct/init", "",
     '{"originalName":"x","fileSize":-1,"fileMd5":"<32位>"}',
     "400", "“文件大小必须大于 0”", ""),
])
add(FIL, "POST /api/files/direct/{uploadId}/complete", [
    ("FIL-DC-01", "P1", "正常", "已 init 并 PUT 完整对象到 MinIO", "POST", "/api/files/direct/{uploadId}/complete", "", "",
     "200", "校验对象存在且大小匹配→生成 FileResponse，任务 SUCCESS", ""),
    ("FIL-DC-02", "P1", "异常", "uploadId 不存在/类型不符", "POST", "/api/files/direct/not-exist/complete", "", "",
     "404", "“上传任务不存在/类型不匹配”", ""),
    ("FIL-DC-03", "P1", "业务规则", "任务非 UPLOADING（已完成/已取消）", "POST", "/api/files/direct/{uploadId}/complete", "", "",
     "409", "“上传任务状态不允许完成”", ""),
    ("FIL-DC-04", "P1", "业务规则", "对象大小与声明不符", "POST", "/api/files/direct/{uploadId}/complete", "", "",
     "409", "“文件大小不匹配”", "未 PUT 或内容错误"),
])
add(FIL, "POST /api/files/multipart/init", [
    ("FIL-MI-01", "P0", "正常", "大文件 md5 未命中", "POST", "/api/files/multipart/init", "",
     '{"originalName":"big.mp4","fileSize":16777216,"fileMd5":"<32位>","partSize":8388608}',
     "200", "返回 uploadId/partCount(=2)/partSize/parts[每个含 uploadUrl、expectedSize]", ""),
    ("FIL-MI-02", "P2", "边界", "partSize 小于 5MB", "POST", "/api/files/multipart/init", "",
     '{"originalName":"x","fileSize":16777216,"fileMd5":"<32位>","partSize":1048576}',
     "200", "partSize 被钳制为最小 5MB", "最小分片 5MB"),
    ("FIL-MI-03", "P1", "业务规则", "分片数超过上限 1000", "POST", "/api/files/multipart/init", "",
     '{"originalName":"x","fileSize":100000000000,"fileMd5":"<32位>","partSize":8388608}',
     "400", "“文件分片数量超过上限”", ""),
    ("FIL-MI-04", "P1", "正常", "md5 命中秒传", "POST", "/api/files/multipart/init", "",
     '{"originalName":"x","fileSize":<size>,"fileMd5":"<已存在>"}',
     "200", "instant=true，不创建上传任务", ""),
    ("FIL-MI-05", "P1", "异常", "无", "POST", "/api/files/multipart/init", "",
     '{"originalName":"x","fileSize":1024,"fileMd5":"bad"}',
     "400", "MD5 格式校验失败", ""),
    ("FIL-MI-06", "P2", "边界", "单分片文件", "POST", "/api/files/multipart/init", "",
     '{"originalName":"x","fileSize":1024,"fileMd5":"<32位>","partSize":8388608}',
     "200", "partCount=1", "最小分片场景"),
])
add(FIL, "GET /api/files/multipart/{uploadId}/parts", [
    ("FIL-MP-01", "P1", "正常", "已 init 并上传部分分片", "GET", "/api/files/multipart/{uploadId}/parts", "", "",
     "200", "返回 partCount 与 uploadedParts（已上传分片序号）", "断点续传依据"),
    ("FIL-MP-02", "P2", "正常", "未上传任何分片", "GET", "/api/files/multipart/{uploadId}/parts", "", "",
     "200", "uploadedParts 为空数组", ""),
    ("FIL-MP-03", "P1", "异常", "uploadId 不存在/类型不符", "GET", "/api/files/multipart/not-exist/parts", "", "",
     "404", "“上传任务不存在/类型不匹配”", ""),
])
add(FIL, "POST /api/files/multipart/{uploadId}/complete", [
    ("FIL-MC-01", "P1", "正常", "全部分片已 PUT", "POST", "/api/files/multipart/{uploadId}/complete", "", "",
     "200", "逐片校验→composeObject→生成 FileResponse，任务 SUCCESS", ""),
    ("FIL-MC-02", "P1", "业务规则", "缺少分片", "POST", "/api/files/multipart/{uploadId}/complete", "", "",
     "409", "分片对象不存在/大小为 0，抛 IllegalStateException", ""),
    ("FIL-MC-03", "P1", "业务规则", "某分片大小与 expected 不符", "POST", "/api/files/multipart/{uploadId}/complete", "", "",
     "409", "“分片大小不匹配”", ""),
    ("FIL-MC-04", "P2", "业务规则", "合并对象总大小与声明不符", "POST", "/api/files/multipart/{uploadId}/complete", "", "",
     "409", "“文件大小不匹配”", ""),
    ("FIL-MC-05", "P1", "业务规则", "任务非 UPLOADING", "POST", "/api/files/multipart/{uploadId}/complete", "", "",
     "409", "“上传任务状态不允许完成”", ""),
    ("FIL-MC-06", "P1", "异常", "uploadId 不存在", "POST", "/api/files/multipart/not-exist/complete", "", "",
     "404", "“上传任务不存在”", ""),
])
add(FIL, "POST /api/files/multipart/{uploadId}/abort", [
    ("FIL-MA-01", "P1", "正常", "已 init 的分片任务", "POST", "/api/files/multipart/{uploadId}/abort", "", "",
     "200", "aborted=true；任务 ABORTED，提交后清理临时分片", ""),
    ("FIL-MA-02", "P2", "业务规则", "任务已 ABORTED/SUCCESS", "POST", "/api/files/multipart/{uploadId}/abort", "", "",
     "409", "非 UPLOADING 抛 IllegalStateException", ""),
    ("FIL-MA-03", "P1", "异常", "uploadId 不存在", "POST", "/api/files/multipart/not-exist/abort", "", "",
     "404", "“上传任务不存在”", ""),
])
add(FIL, "GET /api/files/{fileId}", [
    ("FIL-GET-01", "P0", "正常", "存在 NORMAL 文件", "GET", "/api/files/{fileId}", "", "",
     "200", "返回 FileResponse（originalName/fileSize/fileMd5/fileExt/contentType/createdAt）", ""),
    ("FIL-GET-02", "P1", "异常", "fileId 不存在", "GET", "/api/files/not-exist", "", "",
     "404", "“文件不存在”", ""),
    ("FIL-GET-03", "P2", "业务规则", "文件已逻辑删除", "GET", "/api/files/{fileId}", "", "",
     "404", "DELETED 记录不可查询", ""),
])
add(FIL, "GET /api/files/{fileId}/download", [
    ("FIL-DL-01", "P0", "正常", "存在 NORMAL 文件", "GET", "/api/files/{fileId}/download", "", "",
     "302", "返回 302 + Location（MinIO 签名下载地址）", ""),
    ("FIL-DL-02", "P1", "异常", "fileId 不存在", "GET", "/api/files/not-exist/download", "", "",
     "404", "“文件不存在”", ""),
    ("FIL-DL-03", "P2", "业务规则", "文件已删除", "GET", "/api/files/{fileId}/download", "", "",
     "404", "DELETED 不可下载", ""),
])
add(FIL, "POST /api/files/{fileId}/delete", [
    ("FIL-FDEL-01", "P0", "正常", "存在 NORMAL 文件", "POST", "/api/files/{fileId}/delete", "", "",
     "200", "deleted=true；记录置 DELETED，提交后清理 MinIO 对象", "逻辑删除"),
    ("FIL-FDEL-02", "P1", "异常", "fileId 不存在", "POST", "/api/files/not-exist/delete", "", "",
     "404", "“文件不存在”", ""),
    ("FIL-FDEL-03", "P2", "业务规则", "重复删除同一文件", "POST", "/api/files/{fileId}/delete", "", "",
     "404", "已 DELETED 不再命中 NORMAL → 404", ""),
])
add(FIL, "POST /api/files/page", [
    ("FIL-FP-01", "P0", "正常", "存在 NORMAL 文件", "POST", "/api/files/page", "",
     '{"pageNo":1,"pageSize":10}', "200", "返回 total/pageNo/pageSize/records（仅 NORMAL）", ""),
    ("FIL-FP-02", "P2", "正常", "按名称/扩展名过滤", "POST", "/api/files/page", "",
     '{"originalName":"demo","fileExt":"zip"}', "200", "按 originalName/fileExt 可选过滤", ""),
    ("FIL-FP-03", "P2", "边界", "无", "POST", "/api/files/page", "", "{}",
     "200", "pageNo=1，pageSize 缺省=20", "缺省分页"),
    ("FIL-FP-04", "P2", "边界", "无", "POST", "/api/files/page", "",
     '{"pageNo":0,"pageSize":9999}', "200", "pageNo=1；pageSize 受 max-page-size(100) 限制", ""),
])

# ============================== 接口参数列表 ==============================
PARAMS = [
    ("Excel 导入导出", "查询学生总数", "GET", "/api/excel/count", "—", "—", "—", "{count}", "200", "无", ""),
    ("Excel 导入导出", "生成演示数据", "POST", "/api/excel/seed/{count}", "count:int", "—", "—", "{inserted,count,elapsedMs}", "200/400", "无", ""),
    ("Excel 导入导出", "提交导出任务", "POST", "/api/excel/export", "—", "X-User-Id?", "—", "ExportTaskResponse", "200", "归属", "异步"),
    ("Excel 导入导出", "查询导出任务状态", "GET", "/api/excel/export/{taskId}", "taskId", "X-User-Id?", "—", "ExportTaskResponse", "200/404", "归属", ""),
    ("Excel 导入导出", "下载导出文件", "GET", "/api/excel/export/{taskId}/download", "taskId", "X-User-Id?", "—", "302 Location", "302/404/409", "归属", ""),
    ("Excel 导入导出", "下载导入模板", "GET", "/api/excel/template", "—", "—", "—", "xlsx 流", "200", "无", ""),
    ("Excel 导入导出", "提交导入任务", "POST", "/api/excel/import", "—", "X-User-Id?", "multipart file", "ImportTaskResponse", "200/400/413", "归属", "异步"),
    ("Excel 导入导出", "查询导入任务状态", "GET", "/api/excel/import/{taskId}", "taskId", "X-User-Id?", "—", "ImportTaskResponse", "200/404", "归属", ""),
    ("Excel 导入导出", "下载导入错误明细", "GET", "/api/excel/import/{taskId}/error-file", "taskId", "X-User-Id?", "—", "302 Location", "302/404", "归属", ""),
    ("异步任务中心", "查询任务详情", "GET", "/api/tasks/{taskId}", "taskId", "X-User-Id?", "—", "AsyncTaskResponse", "200/404", "归属", ""),
    ("异步任务中心", "分页查询任务", "POST", "/api/tasks/page", "—", "X-User-Id?", "{pageNo,pageSize,taskType?,status?}", "AsyncTaskPageResponse", "200/400", "归属", ""),
    ("异步任务中心", "取消任务", "POST", "/api/tasks/{taskId}/cancel", "taskId", "X-User-Id?", "—", "{taskId,canceled}", "200/404", "归属", ""),
    ("异步任务中心", "重试任务", "POST", "/api/tasks/{taskId}/retry", "taskId", "X-User-Id?", "—", "AsyncTaskResponse", "200/404/409", "归属", ""),
    ("报表运行控制", "分页查询运行控制", "POST", "/api/report/student-runs/page", "—", "X-User-Id?", "{pageNo,pageSize,runName?,status?}", "StudentReportRunPageResponse", "200/400", "归属", ""),
    ("报表运行控制", "创建运行控制", "POST", "/api/report/student-runs/create", "—", "X-User-Id?", "StudentReportRunCreateRequest", "StudentReportRunResponse", "200/400", "归属", ""),
    ("报表运行控制", "查询运行控制详情", "GET", "/api/report/student-runs/{runId}", "runId", "X-User-Id?", "—", "StudentReportRunResponse", "200/404", "归属", ""),
    ("报表运行控制", "修改运行控制", "POST", "/api/report/student-runs/{runId}/update", "runId", "X-User-Id?", "StudentReportRunUpdateRequest", "StudentReportRunResponse", "200/400/404", "归属", ""),
    ("报表运行控制", "删除运行控制", "POST", "/api/report/student-runs/{runId}/delete", "runId", "X-User-Id?", "—", "{runId,deleted}", "200/404", "归属", "逻辑删除"),
    ("报表运行控制", "运行报表创建导出", "POST", "/api/report/student-runs/{runId}/run", "runId", "X-User-Id?", "—", "ExportTaskResponse", "200/404", "归属", "businessKey=runId"),
    ("报表运行控制", "查询运行历史任务", "POST", "/api/report/student-runs/{runId}/tasks", "runId", "X-User-Id?", "AsyncTaskPageQueryRequest", "AsyncTaskPageResponse", "200/404", "归属", ""),
    ("文件上传中心", "后端上传文件", "POST", "/api/files/upload", "—", "—", "multipart file", "FileUploadResponse", "200/400", "无", ""),
    ("文件上传中心", "秒传检查", "POST", "/api/files/instant-check", "—", "—", "{fileMd5,fileSize}", "InstantUploadCheckResponse", "200/400", "无", ""),
    ("文件上传中心", "初始化直传", "POST", "/api/files/direct/init", "—", "—", "DirectUploadInitRequest", "DirectUploadInitResponse", "200/400", "无", ""),
    ("文件上传中心", "完成直传", "POST", "/api/files/direct/{uploadId}/complete", "uploadId", "—", "—", "FileResponse", "200/404/409", "无", ""),
    ("文件上传中心", "初始化分片上传", "POST", "/api/files/multipart/init", "—", "—", "MultipartUploadInitRequest", "MultipartUploadInitResponse", "200/400", "无", ""),
    ("文件上传中心", "查询分片进度", "GET", "/api/files/multipart/{uploadId}/parts", "uploadId", "—", "—", "MultipartPartsResponse", "200/404", "无", "断点续传"),
    ("文件上传中心", "完成分片上传", "POST", "/api/files/multipart/{uploadId}/complete", "uploadId", "—", "—", "FileResponse", "200/404/409", "无", ""),
    ("文件上传中心", "取消分片上传", "POST", "/api/files/multipart/{uploadId}/abort", "uploadId", "—", "—", "{aborted,uploadId}", "200/404/409", "无", ""),
    ("文件上传中心", "查询文件详情", "GET", "/api/files/{fileId}", "fileId", "—", "—", "FileResponse", "200/404", "无", ""),
    ("文件上传中心", "下载文件", "GET", "/api/files/{fileId}/download", "fileId", "—", "—", "302 Location", "302/404", "无", ""),
    ("文件上传中心", "逻辑删除文件", "POST", "/api/files/{fileId}/delete", "fileId", "—", "—", "{deleted,fileId}", "200/404", "无", ""),
    ("文件上传中心", "分页查询文件", "POST", "/api/files/page", "—", "—", "FilePageQueryRequest", "FilePageResponse", "200", "无", ""),
]

# ============================== 样式 ==============================
thin = Side(style="thin", color="B0B0B0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
cell_font = Font(name="微软雅黑", size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(ws, ncols, fill_color="404040"):
    fill = PatternFill("solid", fgColor=fill_color)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

def apply_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================== 写 测试用例 ==============================
wb = Workbook()
ws = wb.active
ws.title = "测试用例"
for i, (name, width) in enumerate(COLS, 1):
    ws.cell(row=1, column=i, value=name)
style_header(ws, len(COLS))
apply_widths(ws, [w for _, w in COLS])

row = 2
module_rows = {}
for case in CASES:
    (cid, module, endpoint, prio, ctype, pre, method, path,
     headers, body, status, expect, note) = case
    values = [cid, module, endpoint, prio, ctype, pre, method, path, headers, body, status, expect, note]
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = cell_font
        cell.border = border
        cell.alignment = left if c in LEFT_COLS else center
    # 优先级着色（第 4 列）
    pcell = ws.cell(row=row, column=4)
    if prio in PRIO_FILL:
        pcell.fill = PatternFill("solid", fgColor=PRIO_FILL[prio])
        pcell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    # 发现项高亮（第 13 列 备注）
    if note and ("⚠️" in note or "发现项" in note):
        ws.cell(row=row, column=13).fill = PatternFill("solid", fgColor="FFF2CC")
    module_rows.setdefault(module, []).append(row)
    row += 1

# 模块分组配色（第 2 列）
for module, rows in module_rows.items():
    color = MODULE_FILL.get(module, "808080")
    fill = PatternFill("solid", fgColor=color)
    for r in rows:
        c = ws.cell(row=r, column=2)
        c.fill = fill
        c.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        c.alignment = center

ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{row-1}"
ws.sheet_view.showGridLines = False

# ============================== 写 接口参数列表 ==============================
ws2 = wb.create_sheet("接口参数列表")
P_COLS = ["模块", "接口名称", "方法", "路径", "路径参数", "请求头", "请求体/参数", "响应", "主要状态码", "鉴权/归属", "说明"]
P_WIDTHS = [16, 22, 8, 42, 12, 12, 34, 26, 16, 12, 16]
P_LEFT = {4, 7, 8, 11}
for i, name in enumerate(P_COLS, 1):
    ws2.cell(row=1, column=i, value=name)
style_header(ws2, len(P_COLS), fill_color="404040")
apply_widths(ws2, P_WIDTHS)
for r, p in enumerate(PARAMS, 2):
    for c, v in enumerate(p, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.font = cell_font
        cell.border = border
        cell.alignment = left if c in P_LEFT else center
    mod = p[0]
    if mod in MODULE_FILL:
        mc = ws2.cell(row=r, column=1)
        mc.fill = PatternFill("solid", fgColor=MODULE_FILL[mod])
        mc.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        mc.alignment = center
ws2.auto_filter.ref = f"A1:{get_column_letter(len(P_COLS))}{len(PARAMS)+1}"
ws2.sheet_view.showGridLines = False

# ============================== 保存 ==============================
out_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "docs", "test"))
os.makedirs(out_dir, exist_ok=True)
out_path = os.path.join(out_dir, "接口扁平化测试用例.xlsx")
wb.save(out_path)

by_module = {}
for c in CASES:
    by_module[c[1]] = by_module.get(c[1], 0) + 1
print("已生成:", out_path)
print("用例总数: %d  | 接口数: %d" % (len(CASES), len(PARAMS)))
for m, n in by_module.items():
    print("  - %s: %d 条" % (m, n))
