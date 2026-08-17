# -*- coding: utf-8 -*-
"""
全流程接口扁平化测试执行器（Live）。
对运行中的应用 http://localhost:18088 逐接口发起真实请求，记录实际状态码与响应，
覆盖正常 / 边界 / 异常 / 业务规则 / 归属权限分支。

用法（应用已启动）：
    python3 scripts/run_flat_tests.py
产物：
    docs/test/live-test-results.json   结构化结果
    终端打印人类可读的测试记录
"""
import json, os, sys, time, hashlib, subprocess, tempfile, urllib.request

BASE = os.environ.get("BASE_URL", "http://localhost:18088")
OUT_JSON = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "docs", "test", "live-test-results.json"))
TMP = tempfile.mkdtemp(prefix="flat_")
RESULTS = []

def log(msg):
    print(msg, flush=True)

AUTH_TOKENS = {
    "user": os.environ.get("API_SECURITY_DEMO_USER_TOKEN", "").strip(),
    "admin": os.environ.get("API_SECURITY_DEMO_ADMIN_TOKEN", "").strip(),
}

def curl(method, path, headers=None, data=None, form_file=None, upload_file=None,
         raw=False, timeout=180, follow=False, noauth=False):
    url = BASE + path
    headers = dict(headers or {})
    xid = (headers.pop("X-User-Id", "") or "").strip()
    # 鉴权映射：X-User-Id 仅决定身份；user-b/user-cancel 使用 admin 身份做跨归属校验
    if not noauth:
        role = "admin" if xid in ("user-b", "user-cancel") else "user"
        token = AUTH_TOKENS.get(role)
        if token:
            headers["Authorization"] = "Bearer " + token
    # JSON body 请求必须显式声明 Content-Type，否则 Spring @RequestBody 返回 415
    if data is not None and not raw and not form_file:
        if not any(k.lower() == "content-type" for k in headers):
            headers["Content-Type"] = "application/json"
    cmd = ["curl", "-sS", "-X", method,
           "-w", "\n__CODE__%{http_code}\n__LOC__%{redirect_url}"]
    if follow:
        cmd.append("-L")
    for k, v in headers.items():
        cmd += ["-H", "%s: %s" % (k, v)]
    if form_file:
        cmd += ["-F", "file=@%s" % form_file]
    if upload_file is not None:
        cmd += ["--upload-file", upload_file]
    if data is not None:
        cmd += (["--data-binary", data] if raw else ["-d", data])
    cmd += ["-m", str(timeout), url]
    try:
        p = subprocess.run(cmd, capture_output=True, text=False)
        out = p.stdout.decode("utf-8", "replace")
        body, code, loc = out, "", ""
        if "__CODE__" in out:
            parts = out.rsplit("__CODE__", 1)
            body = parts[0]
            tail = parts[1]
            if "__LOC__" in tail:
                code, loc = tail.split("__LOC__", 1)
            else:
                code = tail
        code = (code or "").strip()
        loc = (loc or "").strip()
        err = p.stderr.decode("utf-8", "replace")
        return {"status": code, "body": body, "location": loc, "err": err}
    except Exception as e:
        return {"status": "ERR", "body": "", "location": "", "err": str(e)}

def snippet(s, n=240):
    s = (s or "").replace("\n", " ").strip()
    return s if len(s) <= n else s[:n] + " …"

def P(resp):
    """解析响应体；若为统一响应体 {success,code,message,data} 则返回 data，否则原样返回。"""
    try:
        j = json.loads(resp.get("body", "") or "{}")
    except Exception:
        return {}
    if isinstance(j, dict) and "data" in j and ("success" in j or "code" in j):
        d = j.get("data")
        return d if isinstance(d, dict) else (d if d is not None else {})
    return j if isinstance(j, dict) else {}

def task_status(task_id, owner_id):
    if not task_id:
        return "", {}
    resp = curl("GET", "/api/tasks/%s" % task_id, headers={"X-User-Id": owner_id})
    data = P(resp)
    return data.get("status", ""), data

def expected_retry_status(status):
    # 任务中心只允许 FAILED / CANCELED / EXPIRED 重试；其他状态拒绝重试。
    return "200" if status in ("FAILED", "CANCELED", "EXPIRED") else "409"

def expected_export_download_status(status):
    # 导出成功才会返回 MinIO 预签名 302；失败、取消、超限或仍未完成时应返回 409。
    return "302" if status == "SUCCESS" else "409"

def record(case_id, method, path, expected, resp, extra=None, note=""):
    status = resp.get("status", "")
    exp_codes = [c for c in str(expected).replace("→", "/").split("/") if c.strip()]
    # 异步类（形如 200→任务 FAILED）只校验提交码 200
    verdict = "CHECK"
    if exp_codes:
        if status in exp_codes:
            verdict = "PASS"
        elif status and status not in exp_codes:
            verdict = "FAIL"
    row = {
        "id": case_id, "method": method, "path": path,
        "expected": str(expected), "actual_status": status,
        "verdict": verdict, "body": snippet(resp.get("body", "")),
        "location": resp.get("location", ""), "note": note,
    }
    if extra:
        row.update(extra)
    RESULTS.append(row)
    flag = {"PASS": "✅", "FAIL": "❌", "CHECK": "🔍"}[verdict]
    log("%s [%s] %s %-34s => 期望 %s  实际 %s" % (flag, case_id, method, path, expected, status or resp.get("err","")))
    if resp.get("location"):
        log("        Location: %s" % snippet(resp["location"], 160))
    if resp.get("body"):
        log("        body: %s" % snippet(resp["body"]))
    return resp, row

# ----------------------------- 等待应用 -----------------------------
def wait_up():
    for _ in range(60):
        r = curl("GET", "/api/excel/count", timeout=5)
        if r["status"] == "200":
            return True
        time.sleep(1)
    return False

# ============================== 开始 ==============================
log("=" * 92)
log("全流程接口扁平化测试  BASE=%s  时间=%s" % (BASE, time.strftime("%Y-%m-%d %H:%M:%S")))
log("=" * 92)
if not wait_up():
    log("应用未就绪，终止。"); sys.exit(1)
log("应用就绪。\n")

# -------- 鉴权（云端部署 demo-mode=false） --------
r = curl("GET", "/api/excel/count", noauth=True)
record("SEC-401-01", "GET", "/api/excel/count (no token)", "401", r, note="未携带 Token 应 401")

# -------- Excel 模块 --------
log("──── Excel 导入导出 ────")
r = curl("GET", "/api/excel/count"); record("EXC-COUNT-01", "GET", "/api/excel/count", "200", r)

r = curl("POST", "/api/excel/seed/3"); record("EXC-SEED-01", "POST", "/api/excel/seed/3", "200", r)
r = curl("POST", "/api/excel/seed/0"); record("EXC-SEED-02", "POST", "/api/excel/seed/0", "200", r)
r = curl("POST", "/api/excel/seed/abc"); record("EXC-SEED-03", "POST", "/api/excel/seed/abc", "400", r)

# 下载模板
tpl_path = os.path.join(TMP, "tpl.xlsx")
tpl_cmd = ["curl", "-sS", "-m", "60"]
if AUTH_TOKENS["user"]:
    tpl_cmd += ["-H", "Authorization: Bearer " + AUTH_TOKENS["user"]]
tpl_cmd += ["-o", tpl_path, BASE + "/api/excel/template"]
subprocess.run(tpl_cmd, check=False)
r = curl("GET", "/api/excel/template"); record("EXC-TPL-01", "GET", "/api/excel/template", "200", r,
       extra={"note": "模板已下载 %d 字节" % os.path.getsize(tpl_path)})

# 构造合法导入文件（下载模板 + 追加 3 行新学号）
try:
    from openpyxl import load_workbook
    wb = load_workbook(tpl_path); ws = wb.active
    header = [c.value for c in ws[1]]
    col = {name: i for i, name in enumerate(header)}
    new_rows = [
        ["T20260814001", "张测试", 20, "男", "测试一班", "t001@example.com", "2006-01-01"],
        ["T20260814002", "李测试", 21, "女", "测试一班", "t002@example.com", "2005-05-05"],
        ["T20260814003", "王测试", 22, "男", "测试二班", "t003@example.com", "2004-09-09"],
    ]
    for row in new_rows:
        ws.append(row)
    imp_path = os.path.join(TMP, "import_ok.xlsx"); wb.save(imp_path)
    have_xlsx = True
except Exception as e:
    log("        (构造导入文件失败：%s，跳过部分导入用例)" % e); have_xlsx = False

if have_xlsx:
    r = curl("POST", "/api/excel/import", headers={"X-User-Id": "user-a"}, form_file=imp_path)
    resp, row = record("EXC-IMP-01", "POST", "/api/excel/import", "200", r)
    imp_task_id = ""
    try:
        imp_task_id = P(resp).get("taskId", "")
    except Exception:
        pass
    if imp_task_id:
        time.sleep(2)
        for _ in range(30):
            rs = curl("GET", "/api/excel/import/%s" % imp_task_id, headers={"X-User-Id": "user-a"})
            try:
                st = P(rs).get("status", "")
            except Exception:
                st = ""
            if st in ("SUCCESS", "FAILED"):
                break
            time.sleep(1)
        record("EXC-IMPS-01", "GET", "/api/excel/import/{taskId}", "200", rs, note="导入任务最终状态")
        rb = curl("GET", "/api/excel/import/%s" % imp_task_id, headers={"X-User-Id": "user-b"})
        record("EXC-IMPS-03", "GET", "/api/excel/import/{taskId}", "404", rb)
        rerr = curl("GET", "/api/excel/import/%s/error-file" % imp_task_id, headers={"X-User-Id": "user-a"})
        # 无错误行→404；有错误行→302
        record("EXC-ERRF-02", "GET", "/api/excel/import/{taskId}/error-file", "302/404", rerr)

    # 缺 file 参数
    r = curl("POST", "/api/excel/import", headers={"X-User-Id": "user-a"})
    record("EXC-IMP-03", "POST", "/api/excel/import (no file)", "400/415", r)
    # 空文件
    empty = os.path.join(TMP, "empty.xlsx"); open(empty, "wb").close()
    r = curl("POST", "/api/excel/import", headers={"X-User-Id": "user-a"}, form_file=empty)
    record("EXC-IMP-04", "POST", "/api/excel/import (empty)", "400", r)
    # 非 Excel
    txt = os.path.join(TMP, "notexcel.txt"); open(txt, "w").write("not an excel file")
    r = curl("POST", "/api/excel/import", headers={"X-User-Id": "user-a"}, form_file=txt)
    resp, row = record("EXC-IMP-05", "POST", "/api/excel/import (not excel)", "200", r, note="提交返回 taskId，异步解析应 FAILED")
    bad_task_id = ""
    try: bad_task_id = P(resp).get("taskId", "")
    except Exception: pass
    if bad_task_id:
        time.sleep(2)
        for _ in range(20):
            rs = curl("GET", "/api/excel/import/%s" % bad_task_id, headers={"X-User-Id": "user-a"})
            try: st = P(rs).get("status", "")
            except Exception: st = ""
            if st in ("SUCCESS", "FAILED"): break
            time.sleep(1)
        log("        (非Excel 导入最终状态：%s)" % st)

# 导出（1M 行，需较长时间）
log("        提交导出任务（约 100 万行，需等待）...")
r = curl("POST", "/api/excel/export", headers={"X-User-Id": "user-a"})
resp, row = record("EXC-EXP-01", "POST", "/api/excel/export", "200", r)
exp_task_id = ""
try: exp_task_id = P(resp).get("taskId", "")
except Exception: pass

# 额外提交一个导出任务用于取消测试
r2 = curl("POST", "/api/excel/export", headers={"X-User-Id": "user-cancel"})
cancel_task_id = ""
try: cancel_task_id = P(r2).get("taskId", "")
except Exception: pass
if cancel_task_id:
    rc = curl("POST", "/api/tasks/%s/cancel" % cancel_task_id, headers={"X-User-Id": "user-cancel"})
    record("TSK-CANCEL-01", "POST", "/api/tasks/{taskId}/cancel (active)", "200", rc)

exp_status = ""
if exp_task_id:
    rs = None
    for _ in range(90):
        rs = curl("GET", "/api/excel/export/%s" % exp_task_id, headers={"X-User-Id": "user-a"})
        try: st = P(rs).get("status", "")
        except Exception: st = ""
        if st in ("SUCCESS", "FAILED"): break
        time.sleep(2)
    exp_status = st
    record("EXC-EXPS-01", "GET", "/api/excel/export/{taskId}", "200", rs or curl("GET","/api/excel/export/"+exp_task_id,headers={"X-User-Id":"user-a"}), note="导出最终状态")
    r404 = curl("GET", "/api/excel/not-exist", headers={"X-User-Id": "user-a"})
    record("EXC-EXPS-02", "GET", "/api/excel/not-exist", "404", r404)
    rOwner = curl("GET", "/api/excel/export/%s" % exp_task_id, headers={"X-User-Id": "user-b"})
    record("EXC-EXPS-03", "GET", "/api/excel/export/{taskId} (other owner)", "404", rOwner)
    rd = curl("GET", "/api/excel/export/%s/download" % exp_task_id, headers={"X-User-Id": "user-a"})
    record("EXC-EXPD-01", "GET", "/api/excel/export/{taskId}/download",
           expected_export_download_status(exp_status), rd,
           note="导出任务状态=%s；SUCCESS 才应 302，FAILED/CANCELED/未完成应 409" % (exp_status or "UNKNOWN"))
    rdOwner = curl("GET", "/api/excel/export/%s/download" % exp_task_id, headers={"X-User-Id": "user-b"})
    record("EXC-EXPD-04", "GET", "/api/excel/export/{taskId}/download (other)", "404", rdOwner)

# -------- 任务中心 --------
log("\n──── 异步任务中心 ────")
r = curl("POST", "/api/tasks/page", headers={"X-User-Id": "user-a"}, data='{"pageNo":1,"pageSize":5}')
record("TSK-PAGE-01", "POST", "/api/tasks/page", "200", r)
r = curl("POST", "/api/tasks/page", data='{}')
record("TSK-PAGE-02", "POST", "/api/tasks/page (default)", "200", r)
r = curl("POST", "/api/tasks/page", headers={"X-User-Id": "user-a"}, data='{"pageNo":0,"pageSize":999}')
record("TSK-PAGE-03", "POST", "/api/tasks/page (clamp)", "200", r)
r = curl("POST", "/api/tasks/page", headers={"X-User-Id": "user-a"}, data='{"status":"GARBAGE"}')
record("TSK-PAGE-05", "POST", "/api/tasks/page (bad status)", "400", r, note="统一异常处理返回 400 COMMON_PARAM_ERROR（旧版本为 500，已修复）")
r = curl("POST", "/api/tasks/page", headers={"X-User-Id": "user-a"}, data='{"taskType":"GARBAGE"}')
record("TSK-PAGE-06", "POST", "/api/tasks/page (bad type)", "400", r, note="统一异常处理返回 400")

if exp_task_id:
    r = curl("GET", "/api/tasks/%s" % exp_task_id, headers={"X-User-Id": "user-a"})
    record("TSK-GET-01", "GET", "/api/tasks/{taskId}", "200", r)
    r = curl("GET", "/api/tasks/not-exist", headers={"X-User-Id": "user-a"})
    record("TSK-GET-02", "GET", "/api/tasks/not-exist", "404", r)
    r = curl("GET", "/api/tasks/%s" % exp_task_id, headers={"X-User-Id": "user-b"})
    record("TSK-GET-03", "GET", "/api/tasks/{taskId} (other)", "404", r)
    # 取消已 SUCCESS 的任务 → false
    r = curl("POST", "/api/tasks/%s/cancel" % exp_task_id, headers={"X-User-Id": "user-a"})
    record("TSK-CANCEL-03", "POST", "/api/tasks/{taskId}/cancel (terminal)", "200", r, note="期望 canceled=false")
    # 重试终态任务：FAILED/CANCELED/EXPIRED 允许重试，其余状态拒绝
    exp_retry_status = exp_status
    if not exp_retry_status:
        exp_retry_status, _ = task_status(exp_task_id, "user-a")
    r = curl("POST", "/api/tasks/%s/retry" % exp_task_id, headers={"X-User-Id": "user-a"})
    record("TSK-RETRY-03", "POST", "/api/tasks/{taskId}/retry (terminal)",
           expected_retry_status(exp_retry_status), r,
           note="重试前任务状态=%s；FAILED/CANCELED/EXPIRED 应 200，其余应 409" % (exp_retry_status or "UNKNOWN"))
    r = curl("POST", "/api/tasks/not-exist/retry", headers={"X-User-Id": "user-a"})
    record("TSK-RETRY-06", "POST", "/api/tasks/not-exist/retry", "404", r)

# 重试一个 CANCELED 任务（user-cancel 的导出被取消了）
if cancel_task_id:
    cancel_retry_status, _ = task_status(cancel_task_id, "user-cancel")
    r = curl("POST", "/api/tasks/%s/retry" % cancel_task_id, headers={"X-User-Id": "user-cancel"})
    record("TSK-RETRY-02", "POST", "/api/tasks/{taskId}/retry (CANCELED)",
           expected_retry_status(cancel_retry_status), r,
           note="重试前任务状态=%s；空库或小数据量下取消可能来晚，若已 SUCCESS 则 409 为正确行为" % (cancel_retry_status or "UNKNOWN"))

# -------- 报表运行控制 --------
log("\n──── 报表运行控制 ────")
stamp = "RT%s" % str(int(time.time()))[-8:]
r = curl("POST", "/api/report/student-runs/create", headers={"X-User-Id": "user-a"},
         data='{"runControlCode":"%s","runName":"2026春季测试","studentNo":"T20260814001","minAge":18,"maxAge":25}' % stamp)
resp, row = record("RPT-CRT-01", "POST", "/api/report/student-runs/create", "200", r)
run_id = ""
try: run_id = P(resp).get("runId", "")
except Exception: pass
r = curl("POST", "/api/report/student-runs/create", headers={"X-User-Id": "user-a"}, data='{}')
record("RPT-CRT-02", "POST", "/api/report/student-runs/create (empty)", "400", r)
r = curl("POST", "/api/report/student-runs/create", headers={"X-User-Id": "user-a"},
         data='{"runControlCode":"%s","runName":"dup"}' % stamp)
record("RPT-CRT-03", "POST", "/api/report/student-runs/create (dup code)", "400", r)
r = curl("POST", "/api/report/student-runs/create", headers={"X-User-Id": "user-a"},
         data='{"runControlCode":"%s_b","runName":"x","minAge":-1}' % stamp)
record("RPT-CRT-04", "POST", "/api/report/student-runs/create (age<0)", "400", r)
r = curl("POST", "/api/report/student-runs/create", headers={"X-User-Id": "user-a"},
         data='{"runControlCode":"%s_c","runName":"x","maxAge":200}' % stamp)
record("RPT-CRT-05", "POST", "/api/report/student-runs/create (age>150)", "400", r)
r = curl("POST", "/api/report/student-runs/create", headers={"X-User-Id": "user-a"},
         data='{"runControlCode":"%s_d","runName":"x","minAge":30,"maxAge":20}' % stamp)
record("RPT-CRT-06", "POST", "/api/report/student-runs/create (range)", "400", r)
r = curl("POST", "/api/report/student-runs/page", headers={"X-User-Id": "user-a"}, data='{"pageNo":1,"pageSize":5}')
record("RPT-PAGE-01", "POST", "/api/report/student-runs/page", "200", r)
r = curl("POST", "/api/report/student-runs/page", headers={"X-User-Id": "user-a"}, data='{"status":"GARBAGE"}')
record("RPT-PAGE-04", "POST", "/api/report/student-runs/page (bad status)", "400", r, note="统一异常处理返回 400")

if run_id:
    r = curl("GET", "/api/report/student-runs/%s" % run_id, headers={"X-User-Id": "user-a"})
    record("RPT-GET-01", "GET", "/api/report/student-runs/{runId}", "200", r)
    r = curl("GET", "/api/report/student-runs/not-exist", headers={"X-User-Id": "user-a"})
    record("RPT-GET-02", "GET", "/api/report/student-runs/not-exist", "404", r)
    r = curl("GET", "/api/report/student-runs/%s" % run_id, headers={"X-User-Id": "user-b"})
    record("RPT-GET-03", "GET", "/api/report/student-runs/{runId} (other)", "404", r)
    r = curl("POST", "/api/report/student-runs/%s/update" % run_id, headers={"X-User-Id": "user-a"},
             data='{"runControlCode":"%s","runName":"2026春季改","className":"一班"}' % stamp)
    record("RPT-UPD-01", "POST", "/api/report/student-runs/{runId}/update", "200", r)
    r = curl("POST", "/api/report/student-runs/%s/update" % run_id, headers={"X-User-Id": "user-a"},
             data='{"runControlCode":"%s","runName":"x","minAge":40,"maxAge":30}' % stamp)
    record("RPT-UPD-05", "POST", "/api/report/student-runs/{runId}/update (range)", "400", r)
    # run：用稀有学号过滤，导出快速（几乎 0 行）
    r = curl("POST", "/api/report/student-runs/%s/run" % run_id, headers={"X-User-Id": "user-a"})
    record("RPT-RUN-01", "POST", "/api/report/student-runs/{runId}/run", "200", r)
    r = curl("POST", "/api/report/student-runs/%s/tasks" % run_id, headers={"X-User-Id": "user-a"}, data='{"pageNo":1,"pageSize":5}')
    record("RPT-TASKS-01", "POST", "/api/report/student-runs/{runId}/tasks", "200", r)
    r = curl("POST", "/api/report/student-runs/%s/delete" % run_id, headers={"X-User-Id": "user-a"})
    record("RPT-DEL-01", "POST", "/api/report/student-runs/{runId}/delete", "200", r)
    r = curl("GET", "/api/report/student-runs/%s" % run_id, headers={"X-User-Id": "user-a"})
    record("RPT-GET-04", "GET", "/api/report/student-runs/{runId} (after delete)", "404", r)

# -------- 文件上传中心 --------
log("\n──── 文件上传中心 ────")
demo = os.path.join(TMP, "demo.txt")
content = b"flat-test demo content " + (b"X" * 200)
open(demo, "wb").write(content)
demo_md5 = hashlib.md5(content).hexdigest()
demo_size = len(content)
r = curl("POST", "/api/files/upload", form_file=demo)
resp, row = record("FIL-UP-01", "POST", "/api/files/upload", "200", r)
up_file_id = ""
try: up_file_id = P(resp).get("fileId", "")
except Exception: pass
r = curl("POST", "/api/files/upload")
record("FIL-UP-02", "POST", "/api/files/upload (no file)", "400/415", r)
empty2 = os.path.join(TMP, "empty.txt"); open(empty2, "wb").close()
r = curl("POST", "/api/files/upload", form_file=empty2)
record("FIL-UP-03", "POST", "/api/files/upload (empty)", "400", r)

# 秒传
r = curl("POST", "/api/files/instant-check", data='{"fileMd5":"%s","fileSize":%d}' % (demo_md5, demo_size))
record("FIL-IC-01", "POST", "/api/files/instant-check (hit)", "200", r)
r = curl("POST", "/api/files/instant-check", data='{"fileMd5":"00000000000000000000000000000001","fileSize":%d}' % demo_size)
record("FIL-IC-02", "POST", "/api/files/instant-check (miss)", "200", r)
r = curl("POST", "/api/files/instant-check", data='{"fileMd5":"","fileSize":123}')
record("FIL-IC-03", "POST", "/api/files/instant-check (no md5)", "400", r)
r = curl("POST", "/api/files/instant-check", data='{"fileMd5":"not-hex","fileSize":123}')
record("FIL-IC-04", "POST", "/api/files/instant-check (bad md5)", "400", r)
r = curl("POST", "/api/files/instant-check", data='{"fileMd5":"%s","fileSize":0}' % demo_md5)
record("FIL-IC-05", "POST", "/api/files/instant-check (size 0)", "400", r)

if up_file_id:
    r = curl("GET", "/api/files/%s" % up_file_id)
    record("FIL-GET-01", "GET", "/api/files/{fileId}", "200", r)
    r = curl("GET", "/api/files/not-exist")
    record("FIL-GET-02", "GET", "/api/files/not-exist", "404", r)
    r = curl("POST", "/api/files/page", data='{"pageNo":1,"pageSize":5}')
    record("FIL-FP-01", "POST", "/api/files/page", "200", r)
    r = curl("GET", "/api/files/%s/download" % up_file_id)
    record("FIL-DL-01", "GET", "/api/files/{fileId}/download", "302", r)
    r = curl("GET", "/api/files/not-exist/download")
    record("FIL-DL-02", "GET", "/api/files/not-exist/download", "404", r)

# 直传
r = curl("POST", "/api/files/direct/init",
         data='{"originalName":"demo.zip","contentType":"application/zip","fileSize":%d,"fileMd5":"%s"}' % (demo_size, demo_md5))
resp, row = record("FIL-DI-01", "POST", "/api/files/direct/init", "200", r)
direct_upload_id = direct_url = ""
try:
    j = P(resp); direct_upload_id = j.get("uploadId",""); direct_url = j.get("uploadUrl","")
except Exception: pass
if direct_url:
    rp = subprocess.run(["curl", "-sS", "-X", "PUT", "--data-binary", "@"+demo, "-m", "60", "-w", "%{http_code}", "-o", os.devnull, direct_url],
                        capture_output=True, text=True)
    log("        (直传 PUT 到 MinIO: HTTP %s)" % rp.stdout.strip()[:8])
    r = curl("POST", "/api/files/direct/%s/complete" % direct_upload_id)
    record("FIL-DC-01", "POST", "/api/files/direct/{uploadId}/complete", "200", r)
r = curl("POST", "/api/files/direct/init", data='{"originalName":"x","fileSize":1024,"fileMd5":"bad"}')
record("FIL-DI-04", "POST", "/api/files/direct/init (bad md5)", "400", r)
r = curl("POST", "/api/files/direct/not-exist/complete")
record("FIL-DC-02", "POST", "/api/files/direct/not-exist/complete", "404", r)

# 分片上传（构造 12MB 文件，partSize=5MB → 3 片）
big = os.path.join(TMP, "big.bin")
big_size = 12 * 1024 * 1024
with open(big, "wb") as f:
    f.write(os.urandom(big_size))
big_md5 = hashlib.md5(open(big, "rb").read()).hexdigest()
part_size = 5 * 1024 * 1024
r = curl("POST", "/api/files/multipart/init",
         data='{"originalName":"big.bin","fileSize":%d,"fileMd5":"%s","partSize":%d}' % (big_size, big_md5, part_size))
resp, row = record("FIL-MI-01", "POST", "/api/files/multipart/init", "200", r)
mp_upload_id = ""; parts = []
try:
    j = P(resp); mp_upload_id = j.get("uploadId",""); parts = j.get("parts", [])
except Exception: pass
if parts:
    # 上传第 1、2 片（留第 3 片先不传，测断点）
    for p in parts[:2]:
        num = p["partNumber"]; sz = p["expectedSize"]
        off = (num - 1) * part_size
        chunk = os.path.join(TMP, "part_%d.bin" % num)
        with open(chunk, "wb") as f: f.write(open(big, "rb").read()[off:off + sz])
        rp = subprocess.run(["curl", "-sS", "-X", "PUT", "--upload-file", chunk, "-m", "60", "-w", "%{http_code}", "-o", os.devnull, p["uploadUrl"]],
                            capture_output=True, text=True)
        log("        (分片 %d PUT: HTTP %s)" % (num, rp.stdout.strip()[:8]))
    r = curl("GET", "/api/files/multipart/%s/parts" % mp_upload_id)
    record("FIL-MP-01", "GET", "/api/files/multipart/{uploadId}/parts", "200", r, note="期望 uploadedParts 含已传分片")
    # 上传第 3 片后完成
    p3 = parts[2]; off = (p3["partNumber"] - 1) * part_size
    chunk3 = os.path.join(TMP, "part_3.bin")
    with open(chunk3, "wb") as f: f.write(open(big, "rb").read()[off:off + p3["expectedSize"]])
    subprocess.run(["curl", "-sS", "-X", "PUT", "--upload-file", chunk3, "-m", "60", "-o", os.devnull, p3["uploadUrl"]], capture_output=True)
    r = curl("POST", "/api/files/multipart/%s/complete" % mp_upload_id)
    record("FIL-MC-01", "POST", "/api/files/multipart/{uploadId}/complete", "200", r)
r = curl("GET", "/api/files/multipart/not-exist/parts")
record("FIL-MP-03", "GET", "/api/files/multipart/not-exist/parts", "404", r)

# 取消分片任务
r = curl("POST", "/api/files/multipart/init",
         data='{"originalName":"x.bin","fileSize":%d,"fileMd5":"%s","partSize":%d}' % (big_size, "a"*32, part_size))
resp, _ = record("FIL-MA-init", "POST", "/api/files/multipart/init (for abort)", "200", r)
abort_id = ""
try: abort_id = P(resp).get("uploadId","")
except Exception: pass
if abort_id:
    r = curl("POST", "/api/files/multipart/%s/abort" % abort_id)
    record("FIL-MA-01", "POST", "/api/files/multipart/{uploadId}/abort", "200", r)
    r = curl("POST", "/api/files/multipart/%s/abort" % abort_id)
    record("FIL-MA-02", "POST", "/api/files/multipart/{uploadId}/abort (again)", "409", r)
r = curl("POST", "/api/files/multipart/init", data='{"originalName":"x","fileSize":1024,"fileMd5":"bad"}')
record("FIL-MI-05", "POST", "/api/files/multipart/init (bad md5)", "400", r)

# 逻辑删除
if up_file_id:
    r = curl("POST", "/api/files/%s/delete" % up_file_id)
    record("FIL-FDEL-01", "POST", "/api/files/{fileId}/delete", "200", r)
    r = curl("POST", "/api/files/not-exist/delete")
    record("FIL-FDEL-02", "POST", "/api/files/not-exist/delete", "404", r)
    r = curl("GET", "/api/files/%s" % up_file_id)
    record("FIL-GET-03", "GET", "/api/files/{fileId} (deleted)", "404", r)

# ============================== 汇总 ==============================
from collections import Counter
cnt = Counter(r["verdict"] for r in RESULTS)
log("\n" + "=" * 92)
log("汇总：总用例 %d  ✅PASS %d  ❌FAIL %d  🔍CHECK %d" % (len(RESULTS), cnt.get("PASS",0), cnt.get("FAIL",0), cnt.get("CHECK",0)))
fails = [r for r in RESULTS if r["verdict"] == "FAIL"]
if fails:
    log("失败用例：")
    for r in fails:
        log("  ❌ %s  %s  实际=%s 期望=%s" % (r["id"], r["path"], r["actual_status"], r["expected"]))
log("=" * 92)

os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
with open(OUT_JSON, "w") as f:
    json.dump({"base": BASE, "time": time.strftime("%Y-%m-%d %H:%M:%S"),
               "summary": dict(cnt), "total": len(RESULTS), "results": RESULTS}, f, ensure_ascii=False, indent=2)
log("结构化结果已写入: %s" % OUT_JSON)
